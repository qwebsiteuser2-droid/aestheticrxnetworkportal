#!/usr/bin/env python3
"""
Generate professional PDF and Excel challan/invoice slips for AestheticRXNetwork.
Based on invoice data – logo removed, company name updated, professional format.
"""

import os
from datetime import datetime

# ── Invoice data ──────────────────────────────────────────────────────────────
COMPANY      = "AestheticRXNetwork"
SUBTITLE     = "HAIR CARE AND BEAUTY"
EMAIL        = "atkore.international@gmail.com"
DATE         = "29/1/2026"
BILL_TO      = "Dr Sufyan"
INVOICE_NO   = "X6-5305"

ITEMS = [
    # (Qty, Item #, Description, Unit Price, Total)
    (1.00, "White Medience", "Pdo Cog 100mm", 25_000, 25_000),
]
GRAND_TOTAL  = 25_000

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ═══════════════════════════════════════════════════════════════════════════════
#  PDF Generation (ReportLab)
# ═══════════════════════════════════════════════════════════════════════════════
def generate_pdf(path: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    # ── Colours (matching Excel) ──────────────────────────────────────────────
    BLUE      = HexColor("#1E66FF")
    DARK_BLUE = HexColor("#0837D7")
    GOLD      = HexColor("#C98513")
    LIGHT_BG  = HexColor("#F5F7FA")
    ALT_ROW   = HexColor("#EEF2FF")
    BORDER_C  = HexColor("#CBD5E1")

    # ── Styles (matching Excel fonts/sizes) ───────────────────────────────────
    styles = getSampleStyleSheet()

    company_style = ParagraphStyle(
        "CompanyName", parent=styles["Normal"],
        fontSize=18, textColor=BLUE, fontName="Helvetica-Bold",
        alignment=TA_LEFT, leading=22,
    )
    title_style = ParagraphStyle(
        "InvoiceTitle", parent=styles["Normal"],
        fontSize=24, textColor=DARK_BLUE, fontName="Helvetica-Bold",
        alignment=TA_RIGHT, leading=28,
    )
    subtitle_style = ParagraphStyle(
        "SubTitle", parent=styles["Normal"],
        fontSize=10, textColor=GOLD, fontName="Helvetica-Bold",
        alignment=TA_LEFT,
    )
    normal = ParagraphStyle(
        "NormalCustom", parent=styles["Normal"],
        fontSize=10, textColor=black, fontName="Helvetica",
    )
    normal_right = ParagraphStyle(
        "NormalRight", parent=normal, alignment=TA_RIGHT,
    )
    bold_left = ParagraphStyle(
        "BoldLeft", parent=normal, fontName="Helvetica-Bold",
    )
    bold_right = ParagraphStyle(
        "BoldRight", parent=normal, fontName="Helvetica-Bold",
        alignment=TA_RIGHT,
    )
    footer_style = ParagraphStyle(
        "Footer", parent=styles["Normal"],
        fontSize=9, textColor=HexColor("#666666"),
        alignment=TA_CENTER, fontName="Helvetica-Oblique",
    )
    header_cell = ParagraphStyle(
        "HeaderCell", parent=normal,
        textColor=white, fontName="Helvetica-Bold", fontSize=10,
    )
    header_cell_right = ParagraphStyle(
        "HeaderCellRight", parent=header_cell, alignment=TA_RIGHT,
    )
    total_style = ParagraphStyle(
        "TotalLabel", parent=normal,
        fontSize=11, textColor=DARK_BLUE, fontName="Helvetica-Bold",
        alignment=TA_RIGHT,
    )

    elements = []

    # ── Column widths (5 columns matching Excel: A=8, B=20, C=30, D=16, E=16)
    total_w = doc.width
    cA = total_w * 0.09   # Qty / narrow
    cB = total_w * 0.22   # Item #
    cC = total_w * 0.33   # Description
    cD = total_w * 0.18   # Unit Price / labels
    cE = total_w * 0.18   # Total / values
    col_widths = [cA, cB, cC, cD, cE]

    ROW_H = 18  # standard row height

    # ═══════════════════════════════════════════════════════════════════════════
    # Build the ENTIRE page as one big table (mirrors the Excel grid exactly)
    # ═══════════════════════════════════════════════════════════════════════════
    data = []
    style_cmds = []
    r = 0  # current row index

    # ── Row 0: Company name (cols 0-2) | INVOICE (cols 3-4) ───────────────────
    data.append([
        Paragraph(COMPANY, company_style), "", "",
        Paragraph("INVOICE", title_style), "",
    ])
    style_cmds += [
        ("SPAN", (0, r), (2, r)),
        ("SPAN", (3, r), (4, r)),
        ("VALIGN", (0, r), (-1, r), "MIDDLE"),
    ]
    r += 1

    # ── Row 1: Subtitle (cols 0-2) | Date: | value ───────────────────────────
    data.append([
        Paragraph(SUBTITLE, subtitle_style), "", "",
        Paragraph("Date:", bold_right), Paragraph(DATE, normal_right),
    ])
    style_cmds += [
        ("SPAN", (0, r), (2, r)),
        ("VALIGN", (0, r), (-1, r), "MIDDLE"),
    ]
    r += 1

    # ── Row 2: Email (cols 0-2) | Invoice No: | value ────────────────────────
    data.append([
        Paragraph(f"Email: {EMAIL}", normal), "", "",
        Paragraph("Invoice No:", bold_right), Paragraph(INVOICE_NO, normal_right),
    ])
    style_cmds += [
        ("SPAN", (0, r), (2, r)),
        ("VALIGN", (0, r), (-1, r), "MIDDLE"),
    ]
    r += 1

    # ── Row 3: blank separator ────────────────────────────────────────────────
    data.append(["", "", "", "", ""])
    r += 1

    # ── Row 4: Bill To ────────────────────────────────────────────────────────
    data.append([
        Paragraph("Bill To:", bold_left),
        Paragraph(BILL_TO, normal),
        "", "", "",
    ])
    r += 1

    # ── Row 5: blank separator ────────────────────────────────────────────────
    data.append(["", "", "", "", ""])
    r += 1

    # ── Row 6: Table header ───────────────────────────────────────────────────
    hdr_row = r
    data.append([
        Paragraph("Qty", header_cell),
        Paragraph("Item #", header_cell),
        Paragraph("Description", header_cell),
        Paragraph("Unit Price", header_cell_right),
        Paragraph("Total", header_cell_right),
    ])
    style_cmds += [
        ("BACKGROUND", (0, r), (-1, r), BLUE),
        ("BOX",        (0, r), (-1, r), 1.2, BLUE),
    ]
    r += 1

    # ── Item rows ─────────────────────────────────────────────────────────────
    first_data_row = r
    for qty, item, desc, price, total in ITEMS:
        data.append([
            Paragraph(f"{qty:.2f}", normal),
            Paragraph(item, normal),
            Paragraph(desc, normal),
            Paragraph(f"{price:,.0f}", normal_right),
            Paragraph(f"{total:,.0f}", normal_right),
        ])
        r += 1

    # ── Empty rows (7 rows, matching Excel) ───────────────────────────────────
    for _ in range(7):
        data.append(["", "", "", "", ""])
        r += 1
    last_data_row = r - 1

    # Alternating row fills for data area (rows first_data_row .. last_data_row)
    for i in range(first_data_row, last_data_row + 1):
        if (i - first_data_row) % 2 == 1:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), ALT_ROW))

    # Grid for data area (header row through last data row)
    style_cmds += [
        ("GRID", (0, hdr_row), (-1, last_data_row), 0.5, BORDER_C),
        ("BOX",  (0, hdr_row), (-1, last_data_row), 1.2, BLUE),
    ]

    # ── Total row ─────────────────────────────────────────────────────────────
    total_row = r
    data.append([
        "", "", "",
        Paragraph("Total", total_style),
        Paragraph(f"{GRAND_TOTAL:,.0f}", total_style),
    ])
    style_cmds += [
        ("BACKGROUND", (0, total_row), (-1, total_row), LIGHT_BG),
        ("GRID",       (0, total_row), (-1, total_row), 0.5, BORDER_C),
        ("LINEABOVE",  (0, total_row), (-1, total_row), 1.5, BLUE),
        ("LINEBELOW",  (0, total_row), (-1, total_row), 1.5, BLUE),
    ]
    r += 1

    # ── Blank row ─────────────────────────────────────────────────────────────
    data.append(["", "", "", "", ""])
    r += 1

    # ── Footer row 1: payable to ──────────────────────────────────────────────
    data.append([
        Paragraph(f"Payable to {COMPANY}", footer_style),
        "", "", "", "",
    ])
    style_cmds.append(("SPAN", (0, r), (4, r)))
    r += 1

    # ── Footer row 2: thank you ───────────────────────────────────────────────
    data.append([
        Paragraph("Thank you for your business!", footer_style),
        "", "", "", "",
    ])
    style_cmds.append(("SPAN", (0, r), (4, r)))
    r += 1

    # ── Global table styling ──────────────────────────────────────────────────
    style_cmds += [
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]

    # Row heights
    row_heights = []
    for i in range(r):
        if i == 0:
            row_heights.append(30)   # company / INVOICE row
        elif i == 3 or i == 5:
            row_heights.append(10)   # blank separator
        elif i == total_row + 1:
            row_heights.append(12)   # blank before footer
        else:
            row_heights.append(ROW_H)

    tbl = Table(data, colWidths=col_widths, rowHeights=row_heights)
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    doc.build(elements)
    print(f"  PDF created: {path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  Excel Generation (openpyxl)
# ═══════════════════════════════════════════════════════════════════════════════
def generate_excel(path: str):
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, numbers,
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ── Colours & fonts ───────────────────────────────────────────────────────
    BLUE      = "1E66FF"
    DARK_BLUE = "0837D7"
    GOLD      = "C98513"
    WHITE     = "FFFFFF"
    LIGHT_BG  = "F5F7FA"
    ALT_ROW   = "EEF2FF"
    BORDER_C  = "CBD5E1"

    thin_border  = Side(style="thin", color=BORDER_C)
    med_border   = Side(style="medium", color=BLUE)
    cell_border  = Border(left=thin_border, right=thin_border,
                          top=thin_border, bottom=thin_border)
    header_border = Border(left=med_border, right=med_border,
                           top=med_border, bottom=med_border)

    company_font   = Font(name="Calibri", size=18, bold=True, color=BLUE)
    subtitle_font  = Font(name="Calibri", size=10, bold=True, color=GOLD)
    normal_font    = Font(name="Calibri", size=10)
    bold_font      = Font(name="Calibri", size=10, bold=True)
    title_font     = Font(name="Calibri", size=24, bold=True, color=DARK_BLUE)
    header_font    = Font(name="Calibri", size=10, bold=True, color=WHITE)
    footer_font    = Font(name="Calibri", size=9, italic=True, color="666666")
    total_font     = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)

    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    alt_fill    = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    total_fill  = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {"A": 8, "B": 20, "C": 30, "D": 16, "E": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ── Company name (merged across A-C) ──────────────────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value=COMPANY)
    c.font = company_font
    c.alignment = Alignment(vertical="center")

    # Invoice title on the right side
    ws.merge_cells(f"D{row}:E{row}")
    c = ws.cell(row=row, column=4, value="INVOICE")
    c.font = title_font
    c.alignment = Alignment(horizontal="right", vertical="center")
    row += 1

    # ── Subtitle ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value=SUBTITLE)
    c.font = subtitle_font

    ws.cell(row=row, column=4, value="Date:").font = bold_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5, value=DATE).font = normal_font
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    row += 1

    # ── Email ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value=f"Email: {EMAIL}")
    c.font = normal_font

    ws.cell(row=row, column=4, value="Invoice No:").font = bold_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5, value=INVOICE_NO).font = normal_font
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    row += 1

    # ── Blank separator ───────────────────────────────────────────────────────
    row += 1

    # ── Bill To ───────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Bill To:").font = bold_font
    ws.cell(row=row, column=2, value=BILL_TO).font = normal_font
    row += 2

    # ── Table header ──────────────────────────────────────────────────────────
    headers = ["Qty", "Item #", "Description", "Unit Price", "Total"]
    header_row = row
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = header_border
        c.alignment = Alignment(
            horizontal="right" if col_idx >= 4 else "left",
            vertical="center",
        )
    row += 1

    # ── Item rows ─────────────────────────────────────────────────────────────
    for idx, (qty, item, desc, price, total) in enumerate(ITEMS):
        fill = alt_fill if idx % 2 == 1 else None
        vals = [qty, item, desc, price, total]
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.font = normal_font
            c.border = cell_border
            if col_idx >= 4:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")
            if col_idx == 1:
                c.number_format = '0.00'
            if fill:
                c.fill = fill
        row += 1

    # Add empty rows (like original layout)
    for i in range(7):
        fill = alt_fill if (len(ITEMS) + i) % 2 == 1 else None
        for col_idx in range(1, 6):
            c = ws.cell(row=row, column=col_idx, value="")
            c.border = cell_border
            if fill:
                c.fill = fill
        row += 1

    # ── Grand Total row ───────────────────────────────────────────────────────
    for col_idx in range(1, 6):
        c = ws.cell(row=row, column=col_idx)
        c.fill = total_fill
        c.border = Border(
            left=thin_border, right=thin_border,
            top=Side(style="medium", color=BLUE),
            bottom=Side(style="medium", color=BLUE),
        )
    ws.cell(row=row, column=4, value="Total").font = total_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5, value=GRAND_TOTAL).font = total_font
    ws.cell(row=row, column=5).number_format = '#,##0'
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    row += 2

    # ── Footer ────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=f"Payable to {COMPANY}")
    c.font = footer_font
    c.alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="Thank you for your business!")
    c.font = footer_font
    c.alignment = Alignment(horizontal="center")

    # ── Print setup ───────────────────────────────────────────────────────────
    ws.print_area = f"A1:E{row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(path)
    print(f"  Excel created: {path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    pdf_path   = os.path.join(OUT_DIR, "challan_slip.pdf")
    excel_path = os.path.join(OUT_DIR, "challan_slip.xlsx")

    print("Generating challan slips …")
    generate_pdf(pdf_path)
    generate_excel(excel_path)
    print("\nDone!")

