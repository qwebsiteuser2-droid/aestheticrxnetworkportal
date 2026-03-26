#!/usr/bin/env python3
"""
Convert an .xlsx challan/invoice to a clean image (PNG).
Reads cell values, merges, fonts, fills and borders from openpyxl
and draws them with Pillow — producing a result similar to a screenshot.
"""

import io
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from PIL import Image, ImageDraw, ImageFont

# ── Configuration ─────────────────────────────────────────────────────────────
XLSX_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "aestheticrxnetwork_challan_slip.xlsx",
)
OUT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "aestheticrxnetwork_challan_slip.png",
)

SCALE       = 2          # render at 2× then optionally down-sample for crisp text
CELL_PAD    = 8 * SCALE  # inner cell padding (px)
DPI         = 150

# ── Helpers ───────────────────────────────────────────────────────────────────

def _hex_to_rgb(theme_or_hex, default=(0, 0, 0)):
    """Convert an openpyxl colour value to an (R, G, B) tuple."""
    if theme_or_hex is None:
        return default
    h = str(theme_or_hex).lstrip("#")
    # openpyxl sometimes gives ARGB (8 chars); strip the alpha
    if len(h) == 8:
        h = h[2:]
    if len(h) != 6:
        return default
    try:
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    except ValueError:
        return default


def _get_font(bold=False, italic=False, size=10):
    """Return a PIL TrueType font, falling back gracefully."""
    size_px = int(size * 1.33 * SCALE)  # pt → px, scaled
    candidates = []
    if bold and italic:
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-BoldOblique.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-BoldItalic.ttf",
        ]
    elif bold:
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        ]
    elif italic:
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
        ]
    else:
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        ]
    for path in candidates:
        if os.path.exists(path):
            return ImageFont.truetype(path, size_px)
    return ImageFont.load_default()


# ── Main rendering ────────────────────────────────────────────────────────────

def render_xlsx_to_image(xlsx_path: str, out_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Determine used range
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # ── Compute column widths (in pixels) ─────────────────────────────────────
    # openpyxl width is in "characters"; ≈ 7 px each at normal zoom
    PX_PER_CHAR = 8 * SCALE
    DEFAULT_W   = 10 * PX_PER_CHAR
    col_widths = []
    for ci in range(1, max_col + 1):
        letter = get_column_letter(ci)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width:
            col_widths.append(int(dim.width * PX_PER_CHAR))
        else:
            col_widths.append(DEFAULT_W)

    # ── Compute row heights (in pixels) ───────────────────────────────────────
    DEFAULT_H = 22 * SCALE
    row_heights = []
    for ri in range(1, max_row + 1):
        dim = ws.row_dimensions.get(ri)
        if dim and dim.height:
            row_heights.append(int(dim.height * SCALE))
        else:
            row_heights.append(DEFAULT_H)

    # ── Precompute cumulative positions ───────────────────────────────────────
    MARGIN = 20 * SCALE
    x_pos = [MARGIN]
    for w in col_widths:
        x_pos.append(x_pos[-1] + w)
    y_pos = [MARGIN]
    for h in row_heights:
        y_pos.append(y_pos[-1] + h)

    img_w = x_pos[-1] + MARGIN
    img_h = y_pos[-1] + MARGIN

    # ── Build merged-cells lookup ─────────────────────────────────────────────
    # merged_map[(row,col)] → (min_row, min_col, max_row, max_col) for slave cells
    merged_map = {}       # slave cells → range tuple
    merged_origins = {}   # origin cells → range tuple
    for rng in ws.merged_cells.ranges:
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                if row == rng.min_row and col == rng.min_col:
                    merged_origins[(row, col)] = (
                        rng.min_row, rng.min_col, rng.max_row, rng.max_col,
                    )
                else:
                    merged_map[(row, col)] = (
                        rng.min_row, rng.min_col, rng.max_row, rng.max_col,
                    )

    # ── Create image ──────────────────────────────────────────────────────────
    img = Image.new("RGB", (img_w, img_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    # ── Draw fills ────────────────────────────────────────────────────────────
    for ri in range(1, max_row + 1):
        for ci in range(1, max_col + 1):
            if (ri, ci) in merged_map:
                continue  # skip slave cells
            cell = ws.cell(row=ri, column=ci)
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb and str(fill.fgColor.rgb) not in ("00000000", "0"):
                rgb = _hex_to_rgb(fill.fgColor.rgb)
                if rgb != (0, 0, 0):  # skip black fills (usually means "no fill")
                    # Determine rect (handle merges)
                    if (ri, ci) in merged_origins:
                        _, _, mr, mc = merged_origins[(ri, ci)]
                    else:
                        mr, mc = ri, ci
                    x1 = x_pos[ci - 1]
                    y1 = y_pos[ri - 1]
                    x2 = x_pos[mc]
                    y2 = y_pos[mr]
                    draw.rectangle([x1, y1, x2, y2], fill=rgb)

    # ── Draw borders ──────────────────────────────────────────────────────────
    for ri in range(1, max_row + 1):
        for ci in range(1, max_col + 1):
            if (ri, ci) in merged_map:
                continue
            cell = ws.cell(row=ri, column=ci)
            brd = cell.border
            if not brd:
                continue
            if (ri, ci) in merged_origins:
                _, _, mr, mc = merged_origins[(ri, ci)]
            else:
                mr, mc = ri, ci
            x1 = x_pos[ci - 1]
            y1 = y_pos[ri - 1]
            x2 = x_pos[mc]
            y2 = y_pos[mr]

            def _side_color(side):
                if side and side.style and side.style != "none":
                    c = _hex_to_rgb(getattr(side.color, "rgb", None) if side.color else None,
                                    default=(180, 180, 180))
                    w = 2 * SCALE if side.style == "medium" else 1 * SCALE
                    return c, w
                return None, 0

            for side_attr, coords in [
                ("top",    [(x1, y1), (x2, y1)]),
                ("bottom", [(x1, y2), (x2, y2)]),
                ("left",   [(x1, y1), (x1, y2)]),
                ("right",  [(x2, y1), (x2, y2)]),
            ]:
                color, width = _side_color(getattr(brd, side_attr, None))
                if color:
                    draw.line(coords, fill=color, width=width)

    # ── Draw cell text ────────────────────────────────────────────────────────
    for ri in range(1, max_row + 1):
        for ci in range(1, max_col + 1):
            if (ri, ci) in merged_map:
                continue
            cell = ws.cell(row=ri, column=ci)
            val = cell.value
            if val is None:
                continue

            # Format the value
            from datetime import datetime, date
            if isinstance(val, datetime):
                text = val.strftime("%d/%m/%Y")
            elif isinstance(val, date):
                text = val.strftime("%d/%m/%Y")
            elif isinstance(val, float):
                if val == int(val):
                    text = f"{int(val):,}"
                else:
                    text = f"{val:.2f}"
            elif isinstance(val, int):
                nf = cell.number_format or ""
                if "#,##0" in nf:
                    text = f"{val:,}"
                else:
                    text = str(val)
            else:
                text = str(val)

            # Font
            f = cell.font
            pil_font = _get_font(
                bold=bool(f.bold),
                italic=bool(f.italic),
                size=f.size if f.size else 10,
            )
            text_color = _hex_to_rgb(
                f.color.rgb if (f.color and f.color.rgb) else None,
                default=(0, 0, 0),
            )

            # Cell rectangle (handle merges)
            if (ri, ci) in merged_origins:
                _, _, mr, mc = merged_origins[(ri, ci)]
            else:
                mr, mc = ri, ci
            x1 = x_pos[ci - 1]
            y1 = y_pos[ri - 1]
            x2 = x_pos[mc]
            y2 = y_pos[mr]

            # Measure text
            bbox = draw.textbbox((0, 0), text, font=pil_font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]

            # Horizontal alignment
            align = cell.alignment.horizontal if cell.alignment else None
            if align == "right":
                tx = x2 - CELL_PAD - tw
            elif align == "center":
                tx = x1 + (x2 - x1 - tw) // 2
            else:
                tx = x1 + CELL_PAD

            # Vertical center
            ty = y1 + (y2 - y1 - th) // 2

            draw.text((tx, ty), text, fill=text_color, font=pil_font)

    # ── Draw embedded images (logos etc.) ────────────────────────────────────
    EMU_PER_PX = 9525  # Excel EMU units per pixel
    for xl_img in ws._images:
        anchor = xl_img.anchor
        # Determine pixel position from the anchor
        # OneCellAnchor has _from with col/row; TwoCellAnchor has _from/_to
        if hasattr(anchor, '_from'):
            anc = anchor._from
            col_idx = anc.col  # 0-based
            row_idx = anc.row  # 0-based
            # Offset within cell (EMU → px)
            off_x = int(anc.colOff / EMU_PER_PX) * SCALE if anc.colOff else 0
            off_y = int(anc.rowOff / EMU_PER_PX) * SCALE if anc.rowOff else 0
            ix = x_pos[col_idx] + off_x
            iy = y_pos[row_idx] + off_y
        else:
            ix, iy = MARGIN, MARGIN  # fallback to top-left

        # Read image data from the xlsx image object
        try:
            img_data = xl_img._data()
            pil_logo = Image.open(io.BytesIO(img_data)).convert("RGBA")
        except Exception:
            continue

        # Scale the logo to match SCALE factor
        logo_w = int(xl_img.width * SCALE) if xl_img.width else pil_logo.width * SCALE
        logo_h = int(xl_img.height * SCALE) if xl_img.height else pil_logo.height * SCALE
        pil_logo = pil_logo.resize((logo_w, logo_h), Image.LANCZOS)

        # Paste with alpha mask for transparency
        img.paste(pil_logo, (ix, iy), pil_logo)

    # ── Save ──────────────────────────────────────────────────────────────────
    # Down-sample to half for a sharp result
    final = img.resize((img_w // SCALE, img_h // SCALE), Image.LANCZOS)
    final.save(out_path, dpi=(DPI, DPI))
    print(f"  Image saved: {out_path}  ({final.width}x{final.height} px)")


if __name__ == "__main__":
    print("Rendering Excel → Image …")
    render_xlsx_to_image(XLSX_PATH, OUT_PATH)
    print("Done!")

